# cook your import googleapiclient.discovery
import googleapiclient.errors
import openpyxl

# Set up the API client
api_service_name = "youtube"
api_version = "v3"
api_key = "<your-api-key>"
youtube = googleapiclient.discovery.build(api_service_name, api_version, developerKey=api_key)

# Load the video names from an Excel sheet
wb = openpyxl.load_workbook("YouTube_May_1.xlsx")
ws = wb.active

# Check if the "views" column exists, and create it if it doesn't
if "views" not in [cell.value for cell in ws[1]]:
    ws.cell(row=1, column=2, value="views")

# Batch the requests
batch = youtube.new_batch_http_request()

for row_num, row in enumerate(ws.iter_rows(min_row=2, min_col=1, values_only=True), start=2):
    video_name = row[0]

    # Search for the video by name
    request = youtube.search().list(
        part="id",
        q=video_name,
        type="video",
        maxResults=1
    )
    batch.add(request, callback=lambda response, row_num=row_num: process_response(response, row_num))

# Execute the batch requests
batch.execute()

# Define a callback function to process each response
def process_response(response, row_num):
    if len(response["items"]) == 0:
        print(f"No video found with name: {video_name}")
        return

    video_id = response["items"][0]["id"]["videoId"]

    # Add a request to get the view count for the video
    request = youtube.videos().list(
        part="statistics",
        id=video_id
    )
    batch.add(request, callback=lambda response, row_num=row_num: update_sheet(response, row_num))

# Define a callback function to update the Excel sheet with the view count
def update_sheet(response, row_num):
    if len(response["items"]) == 0:
        print(f"No statistics found for video with ID: {video_id}")
        return

    view_count = response["items"][0]["statistics"]["viewCount"]

    # Update the view count in the Excel sheet
    ws.cell(row=row_num, column=2, value=view_count)

    print(f"Video {video_name} has {view_count} views")
    
# Update the existing Excel sheet
wb.save("YouTube_May_1.xlsx")
dish here
